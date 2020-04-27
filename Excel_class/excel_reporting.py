from openpyxl import Workbook, load_workbook
from datetime import datetime
class Excel():
    def __init__(self):
        self.date = datetime.now().date()
        self.time = datetime.now().time()
        self.wb = Workbook()
        self.ws = self.wb.active
        self.workbook_name = f"Day Reports/({self.date}).xlsx"
        self.just_created = False #// to avoid double check after file creation
        try:
            self.load = load_workbook(self.workbook_name)
        except FileNotFoundError:
            headers = ["link", "name", "date", "time"]
            for x in range(len(headers)):
                self.ws.cell(row=1, column=x+1, value=headers[x]) #// +1 because it must be at least 1
            self.wb.save(self.workbook_name)
            self.just_created = True
      
    def add_data(self, data, words):
        self.load = load_workbook(self.workbook_name)
        self.ns = self.load.get_sheet_by_name("Sheet")
        if len(words) > 0:
            for one in data:
                if isinstance(words, list):
                    for word in words:
                        if word.casefold() in one[0]: #// casefold to ignore capitols
                            data_row = [one[0], one[1], self.date, self.time]
                            self.ns.append(data_row)
                            break
                else:
                    return False
            self.load.save(self.workbook_name)
            return True
        else:
            print("No key words")
            return False